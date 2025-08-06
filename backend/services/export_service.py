"""
Service pro export dat do různých formátů
Autor: GitHub Copilot  
Datum: 6.8.2025
"""

import os
import tempfile
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Any, Optional

# PDF generování
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.pdfgen import canvas
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

# Registrace fontů pro české znaky
try:
    # Pokus o registraci systémových fontů pro Windows
    import platform
    if platform.system() == "Windows":
        # Windows systémové fonty
        pdfmetrics.registerFont(TTFont('Arial-Unicode', 'C:/Windows/Fonts/arial.ttf'))
        pdfmetrics.registerFont(TTFont('Arial-Unicode-Bold', 'C:/Windows/Fonts/arialbd.ttf'))
    else:
        # Linux/Mac fallback
        pdfmetrics.registerFont(TTFont('Arial-Unicode', '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf'))
        pdfmetrics.registerFont(TTFont('Arial-Unicode-Bold', '/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf'))
except Exception as e:
    print(f"⚠️ Nepodařilo se načíst Unicode fonty: {e}")
    print("📝 Používám standardní fonty - háčky a čárky nemusí fungovat správně")
    # Fallback - použij standardní fonty
    NORMAL_FONT = 'Helvetica'
    BOLD_FONT = 'Helvetica-Bold'
else:
    # Unicode fonty úspěšně načteny
    NORMAL_FONT = 'Arial-Unicode'
    BOLD_FONT = 'Arial-Unicode-Bold'

# Excel generování
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.table import Table as ExcelTable, TableStyleInfo

class ExportService:
    """Service pro generování exportů"""
    
    def __init__(self):
        self.temp_dir = Path(tempfile.gettempdir()) / "storage_app_exports"
        self.temp_dir.mkdir(exist_ok=True)
        
    def create_search_pdf(self, results: List[Dict], filters: Dict) -> str:
        """Vytvoří PDF z výsledků vyhledávání"""
        
        # Vytvoř dočasný soubor
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        pdf_path = self.temp_dir / f"search_export_{timestamp}.pdf"
        
        # Vytvoř PDF dokument
        doc = SimpleDocTemplate(str(pdf_path), pagesize=A4)
        styles = getSampleStyleSheet()
        
        # Vlastní styly s Unicode fonty
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            spaceAfter=20,
            alignment=TA_CENTER,
            fontName=BOLD_FONT
        )
        
        gb_style = ParagraphStyle(
            'GBStyle',
            parent=styles['Normal'],
            fontSize=11,
            spaceAfter=4,
            spaceBefore=8,
            fontName=BOLD_FONT
        )
        
        item_style = ParagraphStyle(
            'ItemStyle', 
            parent=styles['Normal'],
            fontSize=10,
            spaceAfter=2,
            leftIndent=20,  # Odsazení pro položky
            fontName=NORMAL_FONT
        )
        
        info_style = ParagraphStyle(
            'InfoStyle',
            parent=styles['Normal'],
            fontSize=9,
            spaceAfter=12,
            fontName=NORMAL_FONT
        )
        
        # Obsah dokumentu
        story = []
        
        # Nadpis
        story.append(Paragraph("📦 Export výsledků vyhledávání", title_style))
        
        # Informace o exportu
        export_info = f"Datum: {datetime.now().strftime('%d.%m.%Y %H:%M')} | Celkem: {len(results)} GB"
        
        # Přidej informace o filtrech
        if any(filters.values()):
            filter_parts = []
            if filters.get('query'):
                filter_parts.append(f"Hledání: {filters['query']}")
            if filters.get('project'):
                filter_parts.append(f"Projekt: {filters['project']}")
            if filters.get('person'):
                filter_parts.append(f"Osoba: {filters['person']}")
            if filters.get('status'):
                filter_parts.append(f"Stav: {filters['status']}")
            
            if filter_parts:
                export_info += f" | Filtry: {' | '.join(filter_parts)}"
        
        story.append(Paragraph(export_info, info_style))
        story.append(Spacer(1, 10))
        
        # Výsledky - jednoduchý seznam
        for i, result in enumerate(results, 1):
            gb = result['gitterbox']
            items = result['items']
            location = result['location']
            shelf = result['shelf']
            position = result['position']
            
            # Hlavní řádek GB
            gb_info = f"#{gb.cislo_gb} - {gb.zodpovedna_osoba} | {position.nazev_pozice} ({location.nazev} - {shelf.nazev}) | {gb.naplnenost_procenta}% | {gb.pocet_polozek} položek"
            
            if gb.datum_zalozeni:
                gb_info += f" | {gb.datum_zalozeni.strftime('%d.%m.%Y')}"
            
            if gb.poznamka:
                gb_info += f" | {gb.poznamka}"
            
            story.append(Paragraph(gb_info, gb_style))
            
            # Položky s odsazením
            if items:
                for item in items:
                    item_parts = []
                    
                    if item.tma_cislo:
                        item_parts.append(f"TMA: {item.tma_cislo}")
                    
                    if item.projekt:
                        item_parts.append(f"Projekt: {item.projekt}")
                    
                    item_parts.append(item.nazev_dilu)
                    
                    if item.popis_mnozstvi:
                        item_parts.append(f"({item.popis_mnozstvi})")
                    
                    # Expirace
                    if item.sledovat_expiraci and item.expiracni_datum:
                        expiry = item.expiracni_datum.strftime('%d.%m.%Y')
                        if item.je_blizko_expirace:
                            expiry += f" ⚠️ {item.dny_do_expirace} dní"
                        item_parts.append(f"Exp: {expiry}")
                    
                    item_text = " • ".join(item_parts)
                    story.append(Paragraph(item_text, item_style))
            else:
                story.append(Paragraph("• Žádné položky", item_style))
        
        # Vygeneruj PDF
        doc.build(story)
        
        return str(pdf_path)
    
    def create_search_excel(self, results: List[Dict], filters: Dict) -> str:
        """Vytvoří Excel z výsledků vyhledávání"""
        
        # Vytvoř dočasný soubor
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_path = self.temp_dir / f"search_export_{timestamp}.xlsx"
        
        # Vytvoř workbook s UTF-8 podporou
        wb = Workbook()
        ws = wb.active
        ws.title = "Výsledky vyhledávání"
        
        # Styly s Unicode podporou
        header_font = Font(bold=True, color="FFFFFF", name="Arial")
        header_fill = PatternFill("solid", fgColor="366092")
        normal_font = Font(name="Arial")  # Arial má lepší podporu českých znaků
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'), 
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Header informace s Unicode fontem
        ws['A1'] = "📦 Export výsledků vyhledávání"
        ws['A1'].font = Font(bold=True, size=16, name="Arial")
        ws['A2'] = f"Datum exportu: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
        ws['A2'].font = normal_font
        ws['A3'] = f"Celkem výsledků: {len(results)}"
        ws['A3'].font = normal_font
        
        # Filtry
        row = 5
        if any(filters.values()):
            ws[f'A{row}'] = "Aplikované filtry:"
            ws[f'A{row}'].font = Font(bold=True, name="Arial")
            row += 1
            
            for key, value in filters.items():
                if value:
                    filter_names = {
                        'query': 'Vyhledávání',
                        'location_id': 'Lokace ID',
                        'project': 'Projekt', 
                        'person': 'Zodpovědná osoba',
                        'status': 'Stav'
                    }
                    cell = ws[f'A{row}']
                    cell.value = f"• {filter_names.get(key, key)}: {value}"
                    cell.font = normal_font
                    row += 1
        
        # Hlavní tabulka začíná
        row += 2
        
        # Headers
        headers = [
            'GB #', 'Zodpovědná osoba', 'Pozice', 'Lokace', 'Regál', 
            'Naplněnost', 'Položek', 'Datum založení', 'TMA', 'Projekt',
            'Název dílu', 'Množství', 'Expirace', 'Dny do exp.', 'Poznámka GB'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        
        # Data
        row += 1
        for result in results:
            gb = result['gitterbox']
            items = result['items']
            location = result['location']
            shelf = result['shelf']
            position = result['position']
            
            if not items:
                # GB bez položek - jeden řádek
                ws.cell(row=row, column=1, value=gb.cislo_gb)
                ws.cell(row=row, column=2, value=gb.zodpovedna_osoba)
                ws.cell(row=row, column=3, value=position.nazev_pozice)
                ws.cell(row=row, column=4, value=location.nazev)
                ws.cell(row=row, column=5, value=shelf.nazev)
                ws.cell(row=row, column=6, value=f"{gb.naplnenost_procenta}%")
                ws.cell(row=row, column=7, value=gb.pocet_polozek)
                ws.cell(row=row, column=8, value=gb.datum_zalozeni.strftime('%d.%m.%Y') if gb.datum_zalozeni else '')
                ws.cell(row=row, column=15, value=gb.poznamka or '')
                
                # Aplikuj border a font
                for col in range(1, 16):
                    cell = ws.cell(row=row, column=col)
                    cell.border = border
                    cell.font = normal_font
                    
                row += 1
            else:
                # GB s položkami - řádek pro každou položku
                first_item = True
                for item in items:
                    if first_item:
                        # První řádek má info o GB
                        ws.cell(row=row, column=1, value=gb.cislo_gb)
                        ws.cell(row=row, column=2, value=gb.zodpovedna_osoba)
                        ws.cell(row=row, column=3, value=position.nazev_pozice)
                        ws.cell(row=row, column=4, value=location.nazev)
                        ws.cell(row=row, column=5, value=shelf.nazev)
                        ws.cell(row=row, column=6, value=f"{gb.naplnenost_procenta}%")
                        ws.cell(row=row, column=7, value=gb.pocet_polozek)
                        ws.cell(row=row, column=8, value=gb.datum_zalozeni.strftime('%d.%m.%Y') if gb.datum_zalozeni else '')
                        ws.cell(row=row, column=15, value=gb.poznamka or '')
                        first_item = False
                    
                    # Info o položce
                    ws.cell(row=row, column=9, value=item.tma_cislo or '')
                    ws.cell(row=row, column=10, value=item.projekt or '')
                    ws.cell(row=row, column=11, value=item.nazev_dilu)
                    ws.cell(row=row, column=12, value=item.popis_mnozstvi or '')
                    
                    if item.sledovat_expiraci and item.expiracni_datum:
                        ws.cell(row=row, column=13, value=item.expiracni_datum.strftime('%d.%m.%Y'))
                        if item.je_blizko_expirace:
                            ws.cell(row=row, column=14, value=item.dny_do_expirace)
                    
                    # Aplikuj border a font
                    for col in range(1, 16):
                        cell = ws.cell(row=row, column=col)
                        cell.border = border
                        cell.font = normal_font
                    
                    row += 1
        
        # Auto-šířka sloupců
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Uložení
        wb.save(excel_path)
        
        return str(excel_path)
